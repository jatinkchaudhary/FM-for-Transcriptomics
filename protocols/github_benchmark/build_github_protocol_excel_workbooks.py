#!/usr/bin/env python3
import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MODELS = [
    "ESM2_PCA512_prior",
    "Txn_Jatin",
    "Txn_Jatin_contextual",
    "BRIDGE",
    "scGPT",
    "Geneformer",
    "BulkFormer_37M",
    "BulkFormer_50M",
    "BulkFormer_93M",
    "BulkFormer_127M",
    "BulkFormer_147M",
]
METRICS = ["AUROC", "AUPRC", "PR@10"]
DATASET_ORDER = {"go": 0, "omim": 1, "go_temporal": 2, "sl": 3, "pombe": 4, "tf": 5, "ng": 6}
SCOPE_ORDER = {"all_genes": 0, "intersect": 1}
OPERATION_ORDER = {"": 0, "sum": 1, "product": 2, "concat": 3}

NAVY = "17324D"
TEAL = "147D92"
GREEN = "D9EAD3"
GOLD = "FFF2CC"
PALE_BLUE = "DDEBF7"
PALE_RED = "F4CCCC"
LIGHT_GRAY = "F2F4F7"
MID_GRAY = "D9E1E8"
WHITE = "FFFFFF"
TEXT = "1F2937"
THIN_GRAY = Side(style="thin", color="D8DEE6")


def benchmark_sort_frame(frame):
    ordered = frame.copy()
    ordered["_dataset"] = ordered["dataset"].map(DATASET_ORDER).fillna(99)
    ordered["_scope"] = ordered["scope"].map(SCOPE_ORDER).fillna(99)
    ordered["_operation"] = ordered["operation"].fillna("").map(OPERATION_ORDER).fillna(99)
    return ordered.sort_values(["_dataset", "_scope", "_operation", "benchmark", "model"])


def ordered_benchmarks(frame):
    ordered = benchmark_sort_frame(frame)
    return ordered.drop_duplicates("benchmark")["benchmark"].tolist()


def pivot_metric(frame, metric):
    pivot = frame.pivot(index="benchmark", columns="model", values=metric)
    return pivot.reindex(index=ordered_benchmarks(frame), columns=MODELS)


def set_title(sheet, title, subtitle, end_column):
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    title_cell = sheet.cell(1, 1, title)
    title_cell.font = Font(size=16, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    subtitle_cell = sheet.cell(2, 1, subtitle)
    subtitle_cell.font = Font(size=10, italic=True, color=TEXT)
    subtitle_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    subtitle_cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 30


def style_header(sheet, row, start_column, end_column):
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(row, column)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GRAY)
    sheet.row_dimensions[row].height = 36


def style_data_grid(sheet, start_row, end_row, start_column, end_column):
    for row in range(start_row, end_row + 1):
        fill = PatternFill("solid", fgColor=WHITE if row % 2 else LIGHT_GRAY)
        for column in range(start_column, end_column + 1):
            cell = sheet.cell(row, column)
            cell.fill = fill
            cell.border = Border(bottom=THIN_GRAY)
            cell.alignment = Alignment(vertical="center", wrap_text=column <= 2)


def highlight_row_maxima(sheet, start_row, end_row, start_column, end_column):
    for row in range(start_row, end_row + 1):
        numeric = []
        for column in range(start_column, end_column + 1):
            value = sheet.cell(row, column).value
            if isinstance(value, (int, float)) and not pd.isna(value):
                numeric.append(float(value))
        if not numeric:
            continue
        maximum = max(numeric)
        for column in range(start_column, end_column + 1):
            cell = sheet.cell(row, column)
            if isinstance(cell.value, (int, float)) and np.isclose(float(cell.value), maximum, rtol=0, atol=1e-12):
                cell.font = Font(bold=True, color=TEXT)
                cell.fill = PatternFill("solid", fgColor=GOLD)


def format_metric_columns(sheet, start_row, end_row, start_column, end_column):
    for row in range(start_row, end_row + 1):
        for column in range(start_column, end_column + 1):
            sheet.cell(row, column).number_format = "0.0000"


def size_columns(sheet, widths):
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def add_metric_sheet(workbook, name, frame, metric, subtitle):
    sheet = workbook.create_sheet(name)
    pivot = pivot_metric(frame, metric)
    end_column = 1 + len(MODELS)
    set_title(sheet, f"{metric} comparison", subtitle, end_column)
    headers = ["Benchmark"] + MODELS
    for column, header in enumerate(headers, 1):
        sheet.cell(4, column, header)
    style_header(sheet, 4, 1, end_column)

    for row_index, (benchmark, values) in enumerate(pivot.iterrows(), 5):
        sheet.cell(row_index, 1, benchmark)
        for column, model in enumerate(MODELS, 2):
            value = values.get(model)
            sheet.cell(row_index, column, None if pd.isna(value) else float(value))

    last_row = 4 + len(pivot)
    style_data_grid(sheet, 5, last_row, 1, end_column)
    format_metric_columns(sheet, 5, last_row, 2, end_column)
    highlight_row_maxima(sheet, 5, last_row, 2, end_column)
    size_columns(sheet, {1: 31, **{column: 20 for column in range(2, end_column + 1)}})
    sheet.freeze_panes = "B5"
    sheet.auto_filter.ref = f"A4:{get_column_letter(end_column)}{last_row}"
    sheet.sheet_view.showGridLines = False
    return sheet


def add_all_metrics_sheet(workbook, name, frame, subtitle):
    sheet = workbook.create_sheet(name)
    end_column = 2 + len(MODELS)
    set_title(sheet, "All benchmark metrics", subtitle, end_column)
    headers = ["Benchmark", "Metric"] + MODELS
    for column, header in enumerate(headers, 1):
        sheet.cell(4, column, header)
    style_header(sheet, 4, 1, end_column)

    row_index = 5
    for benchmark in ordered_benchmarks(frame):
        subset = frame[frame["benchmark"] == benchmark].set_index("model")
        for metric in METRICS:
            sheet.cell(row_index, 1, benchmark)
            sheet.cell(row_index, 2, metric)
            for column, model in enumerate(MODELS, 3):
                value = subset.at[model, metric] if model in subset.index else np.nan
                sheet.cell(row_index, column, None if pd.isna(value) else float(value))
            row_index += 1

    last_row = row_index - 1
    style_data_grid(sheet, 5, last_row, 1, end_column)
    format_metric_columns(sheet, 5, last_row, 3, end_column)
    highlight_row_maxima(sheet, 5, last_row, 3, end_column)
    size_columns(sheet, {1: 31, 2: 12, **{column: 20 for column in range(3, end_column + 1)}})
    sheet.freeze_panes = "C5"
    sheet.auto_filter.ref = f"A4:{get_column_letter(end_column)}{last_row}"
    sheet.sheet_view.showGridLines = False
    return sheet


def winners_for_metric(frame, benchmark, metric):
    subset = frame[frame["benchmark"] == benchmark].set_index("model")[metric].reindex(MODELS)
    maximum = subset.max(skipna=True)
    winners = [model for model, value in subset.items() if not pd.isna(value) and np.isclose(value, maximum, rtol=0, atol=1e-12)]
    return ", ".join(winners), float(maximum)


def add_winners_sheet(workbook, frame):
    sheet = workbook.create_sheet("Winners")
    headers = ["Benchmark", "Family", "AUROC winner", "AUROC", "AUPRC winner", "AUPRC", "PR@10 winner", "PR@10"]
    set_title(sheet, "Per-benchmark winners", "Higher is better. Tied models are all shown.", len(headers))
    for column, header in enumerate(headers, 1):
        sheet.cell(4, column, header)
    style_header(sheet, 4, 1, len(headers))

    for row_index, benchmark in enumerate(ordered_benchmarks(frame), 5):
        subset = frame[frame["benchmark"] == benchmark]
        sheet.cell(row_index, 1, benchmark)
        sheet.cell(row_index, 2, subset.iloc[0]["family"])
        for offset, metric in enumerate(METRICS):
            winner, value = winners_for_metric(frame, benchmark, metric)
            sheet.cell(row_index, 3 + offset * 2, winner)
            sheet.cell(row_index, 4 + offset * 2, value)

    last_row = 4 + frame["benchmark"].nunique()
    style_data_grid(sheet, 5, last_row, 1, len(headers))
    for column in (4, 6, 8):
        format_metric_columns(sheet, 5, last_row, column, column)
        for row in range(5, last_row + 1):
            sheet.cell(row, column).font = Font(bold=True, color=TEXT)
            sheet.cell(row, column).fill = PatternFill("solid", fgColor=GOLD)
    size_columns(sheet, {1: 31, 2: 14, 3: 27, 4: 12, 5: 27, 6: 12, 7: 27, 8: 12})
    sheet.freeze_panes = "C5"
    sheet.auto_filter.ref = f"A4:H{last_row}"
    sheet.sheet_view.showGridLines = False
    return sheet


def model_summary(frame):
    records = []
    benchmarks = frame["benchmark"].unique()
    ranks = {metric: frame.pivot(index="benchmark", columns="model", values=metric).rank(axis=1, ascending=False, method="average") for metric in METRICS}
    for model in MODELS:
        record = {"Model": model, "Benchmarks": len(benchmarks)}
        overall_ranks = []
        for metric in METRICS:
            values = frame[frame["model"] == model][metric]
            metric_ranks = ranks[metric][model]
            wins = int((metric_ranks == 1).sum())
            top_three = int((metric_ranks <= 3).sum())
            record[f"Mean {metric}"] = float(values.mean())
            record[f"{metric} wins"] = wins
            record[f"{metric} top 3"] = top_three
            record[f"Mean {metric} rank"] = float(metric_ranks.mean())
            overall_ranks.append(float(metric_ranks.mean()))
        record["Overall mean rank"] = float(np.mean(overall_ranks))
        records.append(record)
    return pd.DataFrame(records).sort_values("Overall mean rank").reset_index(drop=True)


def add_summary_sheet(workbook, frame, title):
    summary = model_summary(frame)
    sheet = workbook.create_sheet("Model summary")
    set_title(sheet, title, "Means and ranks summarize only the benchmark rows included in this workbook.", len(summary.columns))
    for column, header in enumerate(summary.columns, 1):
        sheet.cell(4, column, header)
    style_header(sheet, 4, 1, len(summary.columns))
    for row_index, row in enumerate(summary.itertuples(index=False), 5):
        for column, value in enumerate(row, 1):
            sheet.cell(row_index, column, value.item() if hasattr(value, "item") else value)

    last_row = 4 + len(summary)
    style_data_grid(sheet, 5, last_row, 1, len(summary.columns))
    for column, header in enumerate(summary.columns, 1):
        if header.startswith("Mean") or header == "Overall mean rank":
            format_metric_columns(sheet, 5, last_row, column, column)
    best_rank = summary["Overall mean rank"].min()
    rank_column = list(summary.columns).index("Overall mean rank") + 1
    for row in range(5, last_row + 1):
        if np.isclose(sheet.cell(row, rank_column).value, best_rank, rtol=0, atol=1e-12):
            sheet.cell(row, rank_column).font = Font(bold=True, color=TEXT)
            sheet.cell(row, rank_column).fill = PatternFill("solid", fgColor=GOLD)
    size_columns(sheet, {1: 24, **{column: 18 for column in range(2, len(summary.columns) + 1)}})
    sheet.freeze_panes = "B5"
    sheet.auto_filter.ref = f"A4:{get_column_letter(len(summary.columns))}{last_row}"
    sheet.sheet_view.showGridLines = False

    mean_auroc_column = list(summary.columns).index("Mean AUROC") + 1
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Mean AUROC by model"
    chart.y_axis.title = "Model"
    chart.x_axis.title = "Mean AUROC"
    data = Reference(sheet, min_col=mean_auroc_column, min_row=4, max_row=last_row)
    categories = Reference(sheet, min_col=1, min_row=5, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7.5
    chart.width = 14
    sheet.add_chart(chart, f"A{last_row + 3}")
    return sheet


def add_long_sheet(workbook, frame):
    sheet = workbook.create_sheet("Long numeric data")
    columns = ["benchmark", "family", "scope", "dataset", "operation", "model", "AUROC", "AUPRC", "PR@10", "terms", "source"]
    ordered = benchmark_sort_frame(frame)[columns]
    set_title(sheet, "Complete numeric source table", "One row per benchmark and model. This sheet preserves full precision.", len(columns))
    for column, header in enumerate(columns, 1):
        sheet.cell(4, column, header)
    style_header(sheet, 4, 1, len(columns))
    for row_index, row in enumerate(ordered.itertuples(index=False), 5):
        for column, value in enumerate(row, 1):
            sheet.cell(row_index, column, None if pd.isna(value) else value)
    last_row = 4 + len(ordered)
    style_data_grid(sheet, 5, last_row, 1, len(columns))
    format_metric_columns(sheet, 5, last_row, 7, 9)
    size_columns(sheet, {1: 31, 2: 14, 3: 14, 4: 14, 5: 14, 6: 24, 7: 12, 8: 12, 9: 12, 10: 10, 11: 65})
    sheet.freeze_panes = "G5"
    sheet.auto_filter.ref = f"A4:K{last_row}"
    sheet.sheet_view.showGridLines = False
    return sheet


def add_provenance_sheet(workbook, frame):
    sheet = workbook.create_sheet("Protocol and notes")
    entries = [
        ("Protocol", "ylaboratory/gene-embedding-benchmarks"),
        ("Pinned commit", "d1320026a2a4ee033d49517f91e2d1c2ccc8df1e"),
        ("Remote run", "github_protocol_d132002_20260721_202549"),
        ("Completion", "2026-07-22T22:20:47Z"),
        ("Models", str(frame["model"].nunique())),
        ("Benchmark groups", str(frame["benchmark"].nunique())),
        ("Model-level results", str(len(frame))),
        ("Failures", "0"),
        ("Included gene-level tasks", "GO, OMIM, Temporal GO"),
        ("Included gene-pair tasks", "SL, POMBE, TF, NG; all_genes/intersect; sum/product/concat"),
        ("Excluded by request", "All ANDES gene-set tasks, including KEGG-GO and disease-tissue"),
        ("AUROC", "Area under the receiver operating characteristic curve; higher is better."),
        ("AUPRC", "Area under the precision-recall curve; higher is better."),
        ("PR@10", "Precision among the ten highest-scored predictions; higher is better."),
        ("Bold cells", "The highest value in each benchmark row is bold and highlighted. Ties are all highlighted."),
        ("Full vs intersect", "all_genes uses each model's available vocabulary; intersect uses the shared gene universe."),
        ("Operations", "sum, product, and concat are the repository's gene-pair embedding operators."),
    ]
    set_title(sheet, "Protocol and interpretation notes", "This workbook reports only the completed GitHub-protocol scope.", 2)
    sheet.cell(4, 1, "Field")
    sheet.cell(4, 2, "Value")
    style_header(sheet, 4, 1, 2)
    for row_index, (field, value) in enumerate(entries, 5):
        sheet.cell(row_index, 1, field)
        sheet.cell(row_index, 2, value)
    last_row = 4 + len(entries)
    style_data_grid(sheet, 5, last_row, 1, 2)
    for row in range(5, last_row + 1):
        sheet.cell(row, 1).font = Font(bold=True, color=TEXT)
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 32
    size_columns(sheet, {1: 25, 2: 105})
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    return sheet


def set_tab_colors(workbook):
    colors = [NAVY, TEAL, "70AD47", "ED7D31", "5B9BD5", "A5A5A5", "FFC000"]
    for index, sheet in enumerate(workbook.worksheets):
        sheet.sheet_properties.tabColor = colors[index % len(colors)]


def save_workbook(workbook, path):
    set_tab_colors(workbook)
    workbook.active = 0
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    load_workbook(path, read_only=True).close()


def build_complete_workbook(frame, path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_summary_sheet(workbook, frame, "All-model benchmark summary")
    add_metric_sheet(workbook, "AUROC comparison", frame, "AUROC", "All 29 completed GitHub-protocol benchmark groups.")
    add_metric_sheet(workbook, "AUPRC comparison", frame, "AUPRC", "All 29 completed GitHub-protocol benchmark groups.")
    add_metric_sheet(workbook, "PR at 10 comparison", frame, "PR@10", "All 29 completed GitHub-protocol benchmark groups.")
    add_all_metrics_sheet(workbook, "All metrics", frame, "All metrics and all models in one table.")
    add_winners_sheet(workbook, frame)
    add_long_sheet(workbook, frame)
    add_provenance_sheet(workbook, frame)
    save_workbook(workbook, path)


def build_gene_level_workbook(frame, path):
    subset = frame[frame["family"] == "gene_level"].copy()
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_summary_sheet(workbook, subset, "Gene-level benchmark summary")
    for metric in METRICS:
        sheet_name = metric if metric != "PR@10" else "PR at 10"
        add_metric_sheet(workbook, sheet_name, subset, metric, "GO, OMIM, and Temporal GO tasks.")
    add_all_metrics_sheet(workbook, "All gene-level metrics", subset, "Five gene-level benchmark groups.")
    add_winners_sheet(workbook, subset)
    add_long_sheet(workbook, subset)
    add_provenance_sheet(workbook, subset)
    save_workbook(workbook, path)


def build_gene_pair_workbook(frame, path):
    subset = frame[frame["family"] == "gene_pair"].copy()
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_summary_sheet(workbook, subset, "Gene-pair benchmark summary")
    for metric in METRICS:
        sheet_name = metric if metric != "PR@10" else "PR at 10"
        add_metric_sheet(workbook, sheet_name, subset, metric, "SL, POMBE, TF, and NG pair tasks.")
    for dataset, label in (("sl", "SL all metrics"), ("pombe", "POMBE all metrics"), ("tf", "TF all metrics"), ("ng", "NG all metrics")):
        add_all_metrics_sheet(workbook, label, subset[subset["dataset"] == dataset], f"{dataset.upper()} full/intersection and three pair operators.")
    add_winners_sheet(workbook, subset)
    add_long_sheet(workbook, subset)
    add_provenance_sheet(workbook, subset)
    save_workbook(workbook, path)


def build_rankings_workbook(frame, path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_summary_sheet(workbook, frame, "Cross-benchmark model rankings")
    add_winners_sheet(workbook, frame)
    add_summary_sheet(workbook, frame[frame["family"] == "gene_level"], "Gene-level model rankings").title = "Gene-level summary"
    add_summary_sheet(workbook, frame[frame["family"] == "gene_pair"], "Gene-pair model rankings").title = "Gene-pair summary"
    add_provenance_sheet(workbook, frame)
    save_workbook(workbook, path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    frame = pd.read_csv(args.input)
    missing_models = sorted(set(MODELS) - set(frame["model"].unique()))
    if missing_models:
        raise RuntimeError(f"Missing models: {missing_models}")
    counts = frame.groupby("benchmark")["model"].nunique()
    if len(counts) != 29 or not (counts == len(MODELS)).all():
        raise RuntimeError(f"Expected 29 complete benchmark groups; got {counts.to_dict()}")

    args.output.mkdir(parents=True, exist_ok=True)
    outputs = [
        args.output / "01_GitHub_Protocol_All_Models.xlsx",
        args.output / "02_Gene_Level_All_Models.xlsx",
        args.output / "03_Gene_Pair_All_Models.xlsx",
        args.output / "04_Model_Rankings_and_Winners.xlsx",
    ]
    build_complete_workbook(frame, outputs[0])
    build_gene_level_workbook(frame, outputs[1])
    build_gene_pair_workbook(frame, outputs[2])
    build_rankings_workbook(frame, outputs[3])
    for output in outputs:
        print(output)


if __name__ == "__main__":
    main()
