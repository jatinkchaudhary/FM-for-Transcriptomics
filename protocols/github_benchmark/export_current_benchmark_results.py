#!/usr/bin/env python3
import argparse
import json
import pickle
from pathlib import Path

import numpy as np
import pandas as pd


MODELS = [
    "ESM2_PCA512_prior",
    "Txn_Jatin",
    "Txn_Jatin_contextual",
    "BRIDGE",
    "scGPT",
    "Geneformer",
    "BulkFormer_37M",
    "BulkFormer_50M",
    "BulkFormer_93M",
    "BulkFormer_127M",
    "BulkFormer_147M",
]
METRICS = ["AUROC", "AUPRC", "PR@10"]
DATASET_LABELS = {"go": "GO", "omim": "OMIM", "ng": "NG", "pombe": "POMBE", "sl": "SL", "tf": "TF"}


def average_pickle_metrics(path):
    with path.open("rb") as handle:
        results = pickle.load(handle)
    frames = [frame for frame in results.values() if isinstance(frame, pd.DataFrame) and not frame.empty]
    if not frames:
        raise ValueError(f"No metric frames in {path}")
    combined = pd.concat(frames, ignore_index=True)
    return {
        "AUROC": float(combined["AUC"].mean()),
        "AUPRC": float(combined["AUPRC"].mean()),
        "PR@10": float(combined["PR@10"].mean()),
        "terms": len(frames),
    }


def model_from_filename(name):
    for model in sorted(MODELS, key=len, reverse=True):
        if name.startswith(model + "_"):
            return model
    return None


def collect_gene_level(results_root):
    rows = []
    configurations = [
        ("all_genes", "go", "GO full"),
        ("intersect", "go", "GO intersect"),
        ("all_genes", "omim", "OMIM full"),
        ("intersect", "omim", "OMIM intersect"),
    ]
    for scope, dataset, benchmark in configurations:
        directory = results_root / "gene_level" / scope / dataset
        for model in MODELS:
            path = directory / f"{model}_holdout_results.pkl"
            if not path.exists():
                continue
            metrics = average_pickle_metrics(path)
            rows.append(
                {
                    "benchmark": benchmark,
                    "family": "gene_level",
                    "scope": scope,
                    "dataset": dataset,
                    "operation": "",
                    "model": model,
                    **metrics,
                    "source": str(path),
                }
            )

    temporal = results_root / "gene_level" / "intersect" / "go_temporal" / "holdout"
    for model in MODELS:
        path = temporal / f"{model}_go_post24_.pkl"
        if not path.exists():
            continue
        metrics = average_pickle_metrics(path)
        rows.append(
            {
                "benchmark": "Temporal GO",
                "family": "gene_level",
                "scope": "intersect",
                "dataset": "go_temporal",
                "operation": "",
                "model": model,
                **metrics,
                "source": str(path),
            }
        )
    return rows


def collect_gene_pair(results_root):
    rows = []
    pair_root = results_root / "gene_pair"
    for path in sorted(pair_root.glob("*/*/*/*.csv")):
        relative = path.relative_to(pair_root)
        scope, dataset, operation = relative.parts[:3]
        model = model_from_filename(path.name)
        if model is None:
            continue
        frame = pd.read_csv(path)
        average = frame[frame["fold"].astype(str).str.lower() == "average"]
        if len(average) != 1:
            continue
        row = average.iloc[0]
        label = DATASET_LABELS.get(dataset, dataset.upper())
        rows.append(
            {
                "benchmark": f"{label} {scope} {operation}",
                "family": "gene_pair",
                "scope": scope,
                "dataset": dataset,
                "operation": operation,
                "model": model,
                "AUROC": float(row["outer_AUC"]),
                "AUPRC": float(row["outer_AUPRC"]),
                "PR@10": float(row["outer_PR@10"]),
                "terms": np.nan,
                "source": str(path),
            }
        )
    return rows


def benchmark_order(name):
    fixed = ["GO full", "GO intersect", "OMIM full", "OMIM intersect", "Temporal GO"]
    if name in fixed:
        return (0, fixed.index(name))
    return (1, name)


def format_number(value):
    return "" if pd.isna(value) else f"{float(value):.4f}"


def build_wide(long_frame, bold):
    records = []
    benchmarks = sorted(long_frame["benchmark"].unique(), key=benchmark_order)
    for benchmark in benchmarks:
        subset = long_frame[long_frame["benchmark"] == benchmark]
        available = subset["model"].nunique()
        for metric in METRICS:
            values = subset.set_index("model")[metric].reindex(MODELS)
            maximum = values.max(skipna=True)
            record = {
                "benchmark": benchmark,
                "metric": metric,
                "models_available": f"{available}/{len(MODELS)}",
                "status": "complete" if available == len(MODELS) else "partial",
            }
            for model, value in values.items():
                rendered = format_number(value)
                if bold and rendered and np.isclose(float(value), float(maximum), rtol=0, atol=1e-12):
                    rendered = f"**{rendered}**"
                record[model] = rendered if bold else value
            records.append(record)
    return pd.DataFrame(records)


def write_excel(path, numeric_wide, long_frame):
    try:
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return False

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        numeric_wide.to_excel(writer, sheet_name="Comparison", index=False)
        long_frame.to_excel(writer, sheet_name="Long numeric", index=False)
        sheet = writer.book["Comparison"]
        model_start = 5
        model_end = model_start + len(MODELS) - 1
        for row in range(2, sheet.max_row + 1):
            values = [sheet.cell(row=row, column=column).value for column in range(model_start, model_end + 1)]
            numeric = [value for value in values if isinstance(value, (int, float))]
            if not numeric:
                continue
            maximum = max(numeric)
            for column in range(model_start, model_end + 1):
                cell = sheet.cell(row=row, column=column)
                if isinstance(cell.value, (int, float)) and np.isclose(cell.value, maximum, rtol=0, atol=1e-12):
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
        sheet.freeze_panes = "E2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 28)
            sheet.column_dimensions[column[0].column_letter].width = width
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--run", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    results_root = args.run / "results"
    args.output.mkdir(parents=True, exist_ok=True)
    rows = collect_gene_level(results_root) + collect_gene_pair(results_root)
    long_frame = pd.DataFrame(rows)
    long_frame = long_frame.sort_values(["family", "benchmark", "model"]).reset_index(drop=True)
    numeric_wide = build_wide(long_frame, bold=False)
    bold_wide = build_wide(long_frame, bold=True)
    auroc_bold = bold_wide[bold_wide["metric"] == "AUROC"].reset_index(drop=True)

    long_frame.to_csv(args.output / "all_results_long_numeric.csv", index=False)
    numeric_wide.to_csv(args.output / "all_results_wide_numeric.csv", index=False)
    bold_wide.to_csv(args.output / "all_results_wide_bold.csv", index=False)
    auroc_bold.to_csv(args.output / "auroc_comparison_bold.csv", index=False)
    excel_created = write_excel(args.output / "all_results_bold.xlsx", numeric_wide, long_frame)

    benchmark_counts = long_frame.groupby("benchmark")["model"].nunique()
    summary = {
        "models": MODELS,
        "model_results": int(len(long_frame)),
        "benchmark_groups": int(len(benchmark_counts)),
        "complete_groups": int((benchmark_counts == len(MODELS)).sum()),
        "partial_groups": {
            benchmark: int(count)
            for benchmark, count in benchmark_counts.items()
            if count != len(MODELS)
        },
        "excel_created": excel_created,
    }
    (args.output / "export_summary.json").write_text(json.dumps(summary, indent=2) + "\n")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
